import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, DoughnutChart
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import re
import tempfile
import os
from typing import Dict, List, Any, Tuple, Optional
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64
import warnings
from datetime import datetime, timedelta
import uuid
import logging

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')

class AdvancedExcelChatbot:
    """
    Advanced Excel Chatbot with comprehensive functionality for:
    - Natural language data querying
    - Data cleaning and transformation
    - Formula generation and debugging
    - Data visualization and charting
    - Pivot table creation
    - Data validation and error checking
    - Automation and scripting
    - Multi-file operations
    """
    
    def __init__(self):
        self.workbook = None
        self.sheets = {}
        self.current_sheet = None
        self.dataframes = {}
        self.charts = {}
        self.pivot_tables = {}
        self.validation_rules = {}
        self.automation_scripts = {}
        self.file_history = []
        self.session_id = str(uuid.uuid4())
        
        # Initialize logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
    def load_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Load Excel file and extract all sheets with comprehensive analysis"""
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.sheets = {}
            self.dataframes = {}
            
            file_info = {
                'filename': os.path.basename(file_path),
                'file_size': os.path.getsize(file_path),
                'sheets_count': len(self.workbook.sheetnames),
                'loaded_at': datetime.now().isoformat()
            }
            
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                # Convert sheet to DataFrame with better handling
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(row)
                
                if data:
                    # Handle headers properly
                    headers = data[0]
                    # Clean headers (remove None values and duplicates)
                    clean_headers = []
                    for i, header in enumerate(headers):
                        if header is None:
                            clean_headers.append(f'Column_{i+1}')
                        else:
                            clean_headers.append(str(header))
                    
                    # Create DataFrame
                    df = pd.DataFrame(data[1:], columns=clean_headers)
                    
                    # Clean data types
                    df = self._infer_and_convert_dtypes(df)
                    
                    self.dataframes[sheet_name] = df
                    self.sheets[sheet_name] = {
                        'dataframe': df,
                        'sheet': sheet,
                        'shape': df.shape,
                        'columns': list(df.columns),
                        'dtypes': df.dtypes.to_dict(),
                        'analysis': self._analyze_sheet(df, sheet_name)
                    }
            
            # Add to file history
            self.file_history.append(file_info)
            
            return {
                'status': 'success',
                'file_info': file_info,
                'sheets': list(self.sheets.keys()),
                'total_sheets': len(self.sheets),
                'summary': self._generate_comprehensive_summary(),
                'recommendations': self._generate_recommendations()
            }
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    def _infer_and_convert_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Intelligently infer and convert data types"""
        for col in df.columns:
            # Try to convert to numeric
            try:
                pd.to_numeric(df[col], errors='raise')
                df[col] = pd.to_numeric(df[col], errors='coerce')
            except:
                # Try to convert to datetime
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                except:
                    # Keep as string
                    df[col] = df[col].astype(str)
        return df
    
    def _analyze_sheet(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Comprehensive analysis of a sheet"""
        analysis = {
            'basic_stats': {
                'rows': len(df),
                'columns': len(df.columns),
                'memory_usage': df.memory_usage(deep=True).sum(),
                'missing_values': df.isnull().sum().sum(),
                'duplicate_rows': df.duplicated().sum()
            },
            'column_analysis': {},
            'data_quality': {},
            'patterns': {}
        }
        
        for col in df.columns:
            col_analysis = {
                'dtype': str(df[col].dtype),
                'missing_count': df[col].isnull().sum(),
                'missing_percentage': (df[col].isnull().sum() / len(df)) * 100,
                'unique_values': df[col].nunique(),
                'duplicate_values': df[col].duplicated().sum()
            }
            
            # Numeric column analysis
            if pd.api.types.is_numeric_dtype(df[col]):
                col_analysis.update({
                    'min': df[col].min(),
                    'max': df[col].max(),
                    'mean': df[col].mean(),
                    'median': df[col].median(),
                    'std': df[col].std(),
                    'outliers': self._detect_outliers(df[col])
                })
            
            # Text column analysis
            elif pd.api.types.is_string_dtype(df[col]):
                col_analysis.update({
                    'avg_length': df[col].str.len().mean(),
                    'max_length': df[col].str.len().max(),
                    'common_values': df[col].value_counts().head(5).to_dict()
                })
            
            # Date column analysis
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                col_analysis.update({
                    'min_date': df[col].min(),
                    'max_date': df[col].max(),
                    'date_range_days': (df[col].max() - df[col].min()).days
                })
            
            analysis['column_analysis'][col] = col_analysis
        
        # Data quality assessment
        analysis['data_quality'] = self._assess_data_quality(df)
        
        # Pattern detection
        analysis['patterns'] = self._detect_patterns(df)
        
        return analysis
    
    def _detect_outliers(self, series: pd.Series) -> Dict[str, Any]:
        """Detect outliers using IQR method"""
        Q1 = series.quantile(0.25)
        Q3 = series.quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR
        
        outliers = series[(series < lower_bound) | (series > upper_bound)]
        
        return {
            'count': len(outliers),
            'percentage': (len(outliers) / len(series)) * 100,
            'values': outliers.tolist()
        }
    
    def _assess_data_quality(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Assess overall data quality"""
        total_cells = df.size
        missing_cells = df.isnull().sum().sum()
        duplicate_rows = df.duplicated().sum()
        
        quality_score = ((total_cells - missing_cells - duplicate_rows) / total_cells) * 100
        
        return {
            'quality_score': quality_score,
            'missing_percentage': (missing_cells / total_cells) * 100,
            'duplicate_percentage': (duplicate_rows / len(df)) * 100,
            'completeness': (1 - (missing_cells / total_cells)) * 100,
            'consistency': self._check_consistency(df)
        }
    
    def _check_consistency(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Check data consistency"""
        consistency_issues = []
        
        # Check for mixed data types in columns
        for col in df.columns:
            if df[col].dtype == 'object':
                # Check if column contains mixed types
                numeric_count = pd.to_numeric(df[col], errors='coerce').notna().sum()
                if 0 < numeric_count < len(df):
                    consistency_issues.append(f"Mixed data types in column '{col}'")
        
        return {
            'issues': consistency_issues,
            'issue_count': len(consistency_issues)
        }
    
    def _detect_patterns(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Detect patterns in the data"""
        patterns = {
            'trends': {},
            'correlations': {},
            'seasonality': {}
        }
        
        # Detect trends in numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            if len(df) > 10:  # Need enough data for trend analysis
                # Simple linear trend
                x = np.arange(len(df))
                y = df[col].fillna(df[col].mean())
                slope = np.polyfit(x, y, 1)[0]
                patterns['trends'][col] = {
                    'slope': slope,
                    'trend': 'increasing' if slope > 0 else 'decreasing' if slope < 0 else 'stable'
                }
        
        # Detect correlations
        if len(numeric_cols) > 1:
            corr_matrix = df[numeric_cols].corr()
            high_corr = []
            for i in range(len(corr_matrix.columns)):
                for j in range(i+1, len(corr_matrix.columns)):
                    corr_val = corr_matrix.iloc[i, j]
                    if abs(corr_val) > 0.7:  # Strong correlation threshold
                        high_corr.append({
                            'col1': corr_matrix.columns[i],
                            'col2': corr_matrix.columns[j],
                            'correlation': corr_val
                        })
            patterns['correlations'] = high_corr
        
        return patterns
    
    def _generate_comprehensive_summary(self) -> Dict[str, Any]:
        """Generate comprehensive summary of all sheets"""
        summary = {
            'overall': {
                'total_sheets': len(self.sheets),
                'total_rows': sum(len(sheet_info['dataframe']) for sheet_info in self.sheets.values()),
                'total_columns': sum(len(sheet_info['dataframe'].columns) for sheet_info in self.sheets.values()),
                'total_memory_usage': sum(sheet_info['dataframe'].memory_usage(deep=True).sum() for sheet_info in self.sheets.values())
            },
            'sheets': {},
            'data_quality_summary': {},
            'recommendations': []
        }
        
        for sheet_name, sheet_info in self.sheets.items():
            df = sheet_info['dataframe']
            analysis = sheet_info['analysis']
            
            summary['sheets'][sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns),
                'data_types': df.dtypes.to_dict(),
                'missing_values': df.isnull().sum().to_dict(),
                'numeric_columns': df.select_dtypes(include=[np.number]).columns.tolist(),
                'text_columns': df.select_dtypes(include=['object']).columns.tolist(),
                'date_columns': df.select_dtypes(include=['datetime64']).columns.tolist(),
                'quality_score': analysis['data_quality']['quality_score'],
                'patterns': analysis['patterns']
            }
        
        # Generate recommendations
        summary['recommendations'] = self._generate_recommendations()
        
        return summary
    
    def _generate_recommendations(self) -> List[str]:
        """Generate intelligent recommendations based on data analysis"""
        recommendations = []
        
        for sheet_name, sheet_info in self.sheets.items():
            df = sheet_info['dataframe']
            analysis = sheet_info['analysis']
            
            # Missing data recommendations
            missing_cols = [col for col, missing in df.isnull().sum().items() if missing > 0]
            if missing_cols:
                recommendations.append(f"Sheet '{sheet_name}': Consider handling missing values in columns: {', '.join(missing_cols)}")
            
            # Data type recommendations
            for col in df.columns:
                if df[col].dtype == 'object':
                    # Check if it should be numeric
                    numeric_count = pd.to_numeric(df[col], errors='coerce').notna().sum()
                    if numeric_count > len(df) * 0.8:  # 80% numeric
                        recommendations.append(f"Sheet '{sheet_name}': Column '{col}' appears to be numeric - consider converting data type")
            
            # Outlier recommendations
            for col in df.select_dtypes(include=[np.number]).columns:
                outliers = analysis['column_analysis'][col].get('outliers', {})
                if outliers.get('count', 0) > 0:
                    recommendations.append(f"Sheet '{sheet_name}': Column '{col}' has {outliers['count']} outliers - consider investigation")
            
            # Correlation recommendations
            correlations = analysis['patterns'].get('correlations', [])
            if correlations:
                recommendations.append(f"Sheet '{sheet_name}': Strong correlations detected - consider visualization")
        
        return recommendations[:10]  # Limit to top 10 recommendations
    
    def natural_language_query(self, query: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        Advanced natural language query processing with intelligent parsing
        Supports complex queries like:
        - "What is the total revenue in Q1?"
        - "Show me the top 5 selling products by region"
        - "Calculate the average profit margin for each department"
        - "Find all transactions above $1000 in March"
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name and sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            # Use first sheet if none specified
            if not sheet_name:
                sheet_name = list(self.sheets.keys())[0]
            
            df = self.sheets[sheet_name]['dataframe']
            
            # Parse the query using advanced NLP techniques
            parsed_query = self._parse_natural_language_query(query, df)
            
            if parsed_query['type'] == 'aggregation':
                return self._execute_aggregation_query(df, parsed_query)
            elif parsed_query['type'] == 'filter':
                return self._execute_filter_query(df, parsed_query)
            elif parsed_query['type'] == 'groupby':
                return self._execute_groupby_query(df, parsed_query)
            elif parsed_query['type'] == 'top_n':
                return self._execute_top_n_query(df, parsed_query)
            elif parsed_query['type'] == 'comparison':
                return self._execute_comparison_query(df, parsed_query)
            elif parsed_query['type'] == 'trend':
                return self._execute_trend_query(df, parsed_query)
            else:
                return self._execute_general_query(df, parsed_query)
                
        except Exception as e:
            self.logger.error(f"Error in natural language query: {str(e)}")
            return {'error': f'Error processing query: {str(e)}'}
    
    def _parse_natural_language_query(self, query: str, df: pd.DataFrame) -> Dict[str, Any]:
        """Advanced natural language query parsing"""
        query_lower = query.lower()
        
        # Define patterns for different query types
        patterns = {
            'aggregation': {
                'keywords': ['sum', 'total', 'average', 'mean', 'count', 'max', 'min', 'highest', 'lowest'],
                'time_patterns': ['q1', 'q2', 'q3', 'q4', 'january', 'february', 'march', 'april', 'may', 'june', 
                                'july', 'august', 'september', 'october', 'november', 'december']
            },
            'filter': {
                'keywords': ['show', 'find', 'filter', 'where', 'above', 'below', 'greater', 'less', 'equal'],
                'comparison_patterns': [r'>\s*(\d+(?:\.\d+)?)', r'<\s*(\d+(?:\.\d+)?)', r'=\s*(\d+(?:\.\d+)?)']
            },
            'groupby': {
                'keywords': ['by', 'group', 'each', 'per', 'for each', 'grouped by']
            },
            'top_n': {
                'keywords': ['top', 'highest', 'best', 'leading'],
                'patterns': [r'top\s+(\d+)', r'(\d+)\s+highest', r'(\d+)\s+best']
            }
        }
        
        # Determine query type
        query_type = 'general'
        for qtype, pattern_info in patterns.items():
            if any(keyword in query_lower for keyword in pattern_info['keywords']):
                query_type = qtype
                break
        
        # Extract columns, values, and conditions
        columns = self._extract_columns_from_query(query, df)
        values = self._extract_values_from_query(query)
        conditions = self._extract_conditions_from_query(query)
        
        return {
            'type': query_type,
            'original_query': query,
            'columns': columns,
            'values': values,
            'conditions': conditions,
            'time_filters': self._extract_time_filters(query),
            'aggregation_type': self._extract_aggregation_type(query)
        }
    
    def _extract_columns_from_query(self, query: str, df: pd.DataFrame) -> List[str]:
        """Extract column names from natural language query"""
        query_lower = query.lower()
        columns = []
        
        # Direct column name matching
        for col in df.columns:
            if col.lower() in query_lower:
                columns.append(col)
        
        # Synonym matching
        column_synonyms = {
            'revenue': ['sales', 'income', 'earnings'],
            'profit': ['margin', 'gain', 'earnings'],
            'cost': ['expense', 'expenditure'],
            'date': ['time', 'period', 'month', 'year'],
            'product': ['item', 'goods', 'merchandise'],
            'customer': ['client', 'buyer', 'user']
        }
        
        for synonym, variants in column_synonyms.items():
            if any(variant in query_lower for variant in variants):
                # Find matching column
                for col in df.columns:
                    if synonym in col.lower():
                        if col not in columns:
                            columns.append(col)
        
        return columns
    
    def _extract_values_from_query(self, query: str) -> List[Any]:
        """Extract numeric values from query"""
        # Extract numbers
        numbers = re.findall(r'\d+(?:\.\d+)?', query)
        return [float(num) if '.' in num else int(num) for num in numbers]
    
    def _extract_conditions_from_query(self, query: str) -> List[Dict[str, Any]]:
        """Extract conditions from query"""
        conditions = []
        
        # Comparison operators
        comparisons = [
            (r'>\s*(\d+(?:\.\d+)?)', '>'),
            (r'<\s*(\d+(?:\.\d+)?)', '<'),
            (r'>=\s*(\d+(?:\.\d+)?)', '>='),
            (r'<=\s*(\d+(?:\.\d+)?)', '<='),
            (r'=\s*(\d+(?:\.\d+)?)', '=')
        ]
        
        for pattern, operator in comparisons:
            matches = re.findall(pattern, query)
            for match in matches:
                conditions.append({
                    'operator': operator,
                    'value': float(match) if '.' in match else int(match)
                })
        
        return conditions
    
    def _extract_time_filters(self, query: str) -> Dict[str, Any]:
        """Extract time-based filters from query"""
        query_lower = query.lower()
        time_filters = {}
        
        # Quarter filters
        quarters = {'q1': 1, 'q2': 2, 'q3': 3, 'q4': 4}
        for quarter, num in quarters.items():
            if quarter in query_lower:
                time_filters['quarter'] = num
        
        # Month filters
        months = {
            'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
            'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
        }
        for month, num in months.items():
            if month in query_lower:
                time_filters['month'] = num
        
        return time_filters
    
    def _extract_aggregation_type(self, query: str) -> str:
        """Extract aggregation type from query"""
        query_lower = query.lower()
        
        if any(word in query_lower for word in ['sum', 'total', 'add']):
            return 'sum'
        elif any(word in query_lower for word in ['average', 'mean', 'avg']):
            return 'mean'
        elif any(word in query_lower for word in ['count', 'number']):
            return 'count'
        elif any(word in query_lower for word in ['max', 'highest', 'maximum']):
            return 'max'
        elif any(word in query_lower for word in ['min', 'lowest', 'minimum']):
            return 'min'
        
        return 'sum'  # Default
    
    def _execute_aggregation_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute aggregation queries"""
        try:
            agg_type = parsed_query['aggregation_type']
            columns = parsed_query['columns']
            
            if not columns:
                # Use first numeric column
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    columns = [numeric_cols[0]]
                else:
                    return {'error': 'No numeric columns found for aggregation'}
            
            # Apply time filters if present
            filtered_df = self._apply_time_filters(df, parsed_query['time_filters'])
            
            # Apply conditions if present
            for condition in parsed_query['conditions']:
                for col in columns:
                    if col in df.columns:
                        if condition['operator'] == '>':
                            filtered_df = filtered_df[filtered_df[col] > condition['value']]
                        elif condition['operator'] == '<':
                            filtered_df = filtered_df[filtered_df[col] < condition['value']]
                        elif condition['operator'] == '>=':
                            filtered_df = filtered_df[filtered_df[col] >= condition['value']]
                        elif condition['operator'] == '<=':
                            filtered_df = filtered_df[filtered_df[col] <= condition['value']]
                        elif condition['operator'] == '=':
                            filtered_df = filtered_df[filtered_df[col] == condition['value']]
            
            # Perform aggregation
            results = {}
            for col in columns:
                if col in filtered_df.columns:
                    if agg_type == 'sum':
                        results[col] = filtered_df[col].sum()
                    elif agg_type == 'mean':
                        results[col] = filtered_df[col].mean()
                    elif agg_type == 'count':
                        results[col] = filtered_df[col].count()
                    elif agg_type == 'max':
                        results[col] = filtered_df[col].max()
                    elif agg_type == 'min':
                        results[col] = filtered_df[col].min()
            
            return {
                'type': 'aggregation',
                'operation': agg_type,
                'columns': columns,
                'results': results,
                'filtered_rows': len(filtered_df),
                'total_rows': len(df),
                'formatted_results': self._format_aggregation_results(results, agg_type)
            }
            
        except Exception as e:
            return {'error': f'Error in aggregation: {str(e)}'}
    
    def _apply_time_filters(self, df: pd.DataFrame, time_filters: Dict[str, Any]) -> pd.DataFrame:
        """Apply time-based filters to DataFrame"""
        if not time_filters:
            return df
        
        # Find date columns
        date_cols = df.select_dtypes(include=['datetime64']).columns
        
        if len(date_cols) == 0:
            return df
        
        # Use first date column
        date_col = date_cols[0]
        
        if 'quarter' in time_filters:
            quarter = time_filters['quarter']
            df = df[df[date_col].dt.quarter == quarter]
        
        if 'month' in time_filters:
            month = time_filters['month']
            df = df[df[date_col].dt.month == month]
        
        return df
    
    def _format_aggregation_results(self, results: Dict[str, Any], agg_type: str) -> str:
        """Format aggregation results for display"""
        formatted = []
        for col, value in results.items():
            if isinstance(value, (int, float)):
                if agg_type in ['sum', 'mean']:
                    formatted.append(f"{agg_type.title()} of {col}: {value:,.2f}")
                else:
                    formatted.append(f"{agg_type.title()} of {col}: {value}")
            else:
                formatted.append(f"{agg_type.title()} of {col}: {value}")
        
        return "; ".join(formatted)
    
    def _execute_filter_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute filter queries"""
        try:
            filtered_df = df.copy()
            
            # Apply conditions
            for condition in parsed_query['conditions']:
                for col in parsed_query['columns']:
                    if col in filtered_df.columns:
                        if condition['operator'] == '>':
                            filtered_df = filtered_df[filtered_df[col] > condition['value']]
                        elif condition['operator'] == '<':
                            filtered_df = filtered_df[filtered_df[col] < condition['value']]
                        elif condition['operator'] == '>=':
                            filtered_df = filtered_df[filtered_df[col] >= condition['value']]
                        elif condition['operator'] == '<=':
                            filtered_df = filtered_df[filtered_df[col] <= condition['value']]
                        elif condition['operator'] == '=':
                            filtered_df = filtered_df[filtered_df[col] == condition['value']]
            
            return {
                'type': 'filter',
                'filtered_rows': len(filtered_df),
                'total_rows': len(df),
                'data': filtered_df.head(20).to_dict('records'),
                'columns': list(filtered_df.columns),
                'summary': {
                    'rows_returned': len(filtered_df),
                    'percentage_of_total': (len(filtered_df) / len(df)) * 100
                }
            }
            
        except Exception as e:
            return {'error': f'Error in filtering: {str(e)}'}
    
    def _execute_groupby_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute groupby queries"""
        try:
            # Identify grouping columns and aggregation columns
            grouping_cols = []
            agg_cols = []
            
            for col in parsed_query['columns']:
                if col in df.columns:
                    if df[col].dtype == 'object':  # Categorical columns for grouping
                        grouping_cols.append(col)
                    else:  # Numeric columns for aggregation
                        agg_cols.append(col)
            
            if not grouping_cols:
                return {'error': 'No grouping columns found'}
            
            if not agg_cols:
                # Use first numeric column for aggregation
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    agg_cols = [numeric_cols[0]]
                else:
                    return {'error': 'No numeric columns found for aggregation'}
            
            # Perform groupby
            grouped = df.groupby(grouping_cols)[agg_cols].agg(['sum', 'mean', 'count']).reset_index()
            
            return {
                'type': 'groupby',
                'grouping_columns': grouping_cols,
                'aggregation_columns': agg_cols,
                'data': grouped.head(20).to_dict('records'),
                'total_groups': len(grouped),
                'summary': f"Grouped by {', '.join(grouping_cols)} with aggregations on {', '.join(agg_cols)}"
            }
            
        except Exception as e:
            return {'error': f'Error in groupby: {str(e)}'}
    
    def _execute_top_n_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute top N queries"""
        try:
            # Extract N from query
            n = 5  # Default
            for value in parsed_query['values']:
                if isinstance(value, int) and 1 <= value <= 100:
                    n = value
                    break
            
            # Find column to sort by
            sort_col = None
            for col in parsed_query['columns']:
                if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                    sort_col = col
                    break
            
            if not sort_col:
                # Use first numeric column
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    sort_col = numeric_cols[0]
                else:
                    return {'error': 'No numeric columns found for sorting'}
            
            # Get top N
            top_n = df.nlargest(n, sort_col)
            
            return {
                'type': 'top_n',
                'n': n,
                'sort_column': sort_col,
                'data': top_n.to_dict('records'),
                'summary': f"Top {n} records by {sort_col}"
            }
            
        except Exception as e:
            return {'error': f'Error in top N query: {str(e)}'}
    
    def _execute_comparison_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute comparison queries"""
        try:
            # This would implement comparison logic
            # For now, return a placeholder
            return {
                'type': 'comparison',
                'message': 'Comparison queries not yet implemented'
            }
        except Exception as e:
            return {'error': f'Error in comparison: {str(e)}'}
    
    def _execute_trend_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute trend analysis queries"""
        try:
            # This would implement trend analysis
            # For now, return a placeholder
            return {
                'type': 'trend',
                'message': 'Trend analysis not yet implemented'
            }
        except Exception as e:
            return {'error': f'Error in trend analysis: {str(e)}'}
    
    def _execute_general_query(self, df: pd.DataFrame, parsed_query: Dict[str, Any]) -> Dict[str, Any]:
        """Execute general queries"""
        try:
            return {
                'type': 'general',
                'message': f"Processed query: {parsed_query['original_query']}",
                'data_preview': df.head(10).to_dict('records'),
                'summary': {
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'columns': list(df.columns)
                }
            }
        except Exception as e:
            return {'error': f'Error in general query: {str(e)}'}

    def generate_formula(self, request: str) -> Dict[str, Any]:
        """
        Generate Excel formulas from natural language requests
        Examples:
        - "Write a formula to calculate the profit margin"
        - "How do I use VLOOKUP to find customer data?"
        - "Create a formula to sum values if date is in Q1"
        """
        try:
            request_lower = request.lower()
            
            # Common formula patterns
            formula_patterns = {
                'sum': {
                    'keywords': ['sum', 'total', 'add', 'calculate total'],
                    'examples': [
                        '=SUM(A2:A100)',
                        '=SUMIF(A2:A100, ">0", B2:B100)',
                        '=SUMIFS(C2:C100, A2:A100, "Product A", B2:B100, ">100")'
                    ]
                },
                'average': {
                    'keywords': ['average', 'mean', 'avg'],
                    'examples': [
                        '=AVERAGE(A2:A100)',
                        '=AVERAGEIF(A2:A100, ">0")',
                        '=AVERAGEIFS(C2:C100, A2:A100, "Product A")'
                    ]
                },
                'count': {
                    'keywords': ['count', 'number of', 'how many'],
                    'examples': [
                        '=COUNT(A2:A100)',
                        '=COUNTIF(A2:A100, "Product A")',
                        '=COUNTIFS(A2:A100, "Product A", B2:B100, ">100")'
                    ]
                },
                'lookup': {
                    'keywords': ['lookup', 'find', 'vlookup', 'hlookup', 'index match'],
                    'examples': [
                        '=VLOOKUP(A2, B2:D100, 3, FALSE)',
                        '=INDEX(C2:C100, MATCH(A2, B2:B100, 0))',
                        '=HLOOKUP(A2, B2:D10, 3, FALSE)'
                    ]
                },
                'conditional': {
                    'keywords': ['if', 'conditional', 'when', 'based on'],
                    'examples': [
                        '=IF(A2>100, "High", "Low")',
                        '=IFS(A2>100, "High", A2>50, "Medium", TRUE, "Low")',
                        '=SWITCH(A2, 1, "One", 2, "Two", "Other")'
                    ]
                },
                'date': {
                    'keywords': ['date', 'time', 'month', 'year', 'quarter'],
                    'examples': [
                        '=MONTH(A2)',
                        '=YEAR(A2)',
                        '=QUARTER(A2)',
                        '=DATEDIF(A2, B2, "D")',
                        '=EOMONTH(A2, 0)'
                    ]
                },
                'text': {
                    'keywords': ['text', 'string', 'concatenate', 'join'],
                    'examples': [
                        '=CONCATENATE(A2, " ", B2)',
                        '=A2&" "&B2',
                        '=LEFT(A2, 5)',
                        '=RIGHT(A2, 3)',
                        '=MID(A2, 2, 4)',
                        '=UPPER(A2)',
                        '=LOWER(A2)',
                        '=PROPER(A2)'
                    ]
                },
                'financial': {
                    'keywords': ['profit', 'margin', 'percentage', 'rate', 'interest'],
                    'examples': [
                        '=A2/B2',
                        '=(A2-B2)/A2',
                        '=PMT(rate/12, nper*12, pv)',
                        '=FV(rate, nper, pmt, pv)',
                        '=PV(rate, nper, pmt, fv)'
                    ]
                }
            }
            
            # Find matching formula type
            matched_type = None
            for formula_type, pattern in formula_patterns.items():
                if any(keyword in request_lower for keyword in pattern['keywords']):
                    matched_type = formula_type
                    break
            
            if not matched_type:
                matched_type = 'general'
            
            # Generate appropriate formula
            if matched_type == 'sum':
                if 'if' in request_lower:
                    formula = '=SUMIF(range, criteria, sum_range)'
                    explanation = "SUMIF adds values in a range that meet a single criterion"
                elif 'ifs' in request_lower:
                    formula = '=SUMIFS(sum_range, criteria_range1, criteria1, criteria_range2, criteria2)'
                    explanation = "SUMIFS adds values in a range that meet multiple criteria"
                else:
                    formula = '=SUM(range)'
                    explanation = "SUM adds all numbers in a range"
            
            elif matched_type == 'average':
                if 'if' in request_lower:
                    formula = '=AVERAGEIF(range, criteria)'
                    explanation = "AVERAGEIF calculates the average of values that meet a criterion"
                else:
                    formula = '=AVERAGE(range)'
                    explanation = "AVERAGE calculates the arithmetic mean of values"
            
            elif matched_type == 'lookup':
                if 'vlookup' in request_lower:
                    formula = '=VLOOKUP(lookup_value, table_array, col_index_num, [range_lookup])'
                    explanation = "VLOOKUP searches for a value in the first column of a table and returns a value in the same row from another column"
                elif 'index' in request_lower and 'match' in request_lower:
                    formula = '=INDEX(return_range, MATCH(lookup_value, lookup_range, 0))'
                    explanation = "INDEX-MATCH is more flexible than VLOOKUP and can search in any column"
                else:
                    formula = '=VLOOKUP(lookup_value, table_array, col_index_num, FALSE)'
                    explanation = "VLOOKUP for exact match (recommended)"
            
            elif matched_type == 'conditional':
                if 'multiple' in request_lower or 'several' in request_lower:
                    formula = '=IFS(condition1, result1, condition2, result2, ...)'
                    explanation = "IFS evaluates multiple conditions and returns the first TRUE result"
                else:
                    formula = '=IF(condition, value_if_true, value_if_false)'
                    explanation = "IF performs a logical test and returns one value if TRUE, another if FALSE"
            
            elif matched_type == 'date':
                if 'month' in request_lower:
                    formula = '=MONTH(date)'
                    explanation = "MONTH extracts the month number (1-12) from a date"
                elif 'year' in request_lower:
                    formula = '=YEAR(date)'
                    explanation = "YEAR extracts the year from a date"
                elif 'quarter' in request_lower:
                    formula = '=ROUNDUP(MONTH(date)/3, 0)'
                    explanation = "Calculates quarter (1-4) from a date"
                else:
                    formula = '=TODAY()'
                    explanation = "TODAY returns the current date"
            
            elif matched_type == 'text':
                if 'join' in request_lower or 'concatenate' in request_lower:
                    formula = '=CONCATENATE(text1, text2, ...) or =A1&" "&B1'
                    explanation = "Concatenates multiple text strings into one"
                elif 'upper' in request_lower:
                    formula = '=UPPER(text)'
                    explanation = "Converts text to uppercase"
                elif 'lower' in request_lower:
                    formula = '=LOWER(text)'
                    explanation = "Converts text to lowercase"
                else:
                    formula = '=LEFT(text, num_chars)'
                    explanation = "LEFT extracts characters from the beginning of a text string"
            
            elif matched_type == 'financial':
                if 'margin' in request_lower:
                    formula = '=(revenue - cost) / revenue'
                    explanation = "Profit margin = (Revenue - Cost) / Revenue"
                elif 'percentage' in request_lower:
                    formula = '=part / total'
                    explanation = "Percentage = Part / Total"
                else:
                    formula = '=A2/B2'
                    explanation = "Basic division for financial calculations"
            
            else:
                formula = '=SUM(A2:A100)'
                explanation = "Basic SUM formula - modify as needed"
            
            return {
                'type': 'formula',
                'formula': formula,
                'explanation': explanation,
                'category': matched_type,
                'examples': formula_patterns.get(matched_type, {}).get('examples', []),
                'usage_tips': self._get_formula_usage_tips(matched_type)
            }
            
        except Exception as e:
            return {'error': f'Error generating formula: {str(e)}'}
    
    def _get_formula_usage_tips(self, formula_type: str) -> List[str]:
        """Get usage tips for different formula types"""
        tips = {
            'sum': [
                "Use SUMIF for conditional summing",
                "Use SUMIFS for multiple conditions",
                "SUM ignores text and empty cells"
            ],
            'average': [
                "AVERAGE ignores empty cells",
                "Use AVERAGEIF for conditional averaging",
                "Consider using MEDIAN for skewed data"
            ],
            'lookup': [
                "VLOOKUP searches left to right only",
                "INDEX-MATCH is more flexible",
                "Always use FALSE for exact matches in VLOOKUP"
            ],
            'conditional': [
                "Nest IF functions for complex logic",
                "Use IFS for multiple conditions (Excel 2019+)",
                "Consider SWITCH for multiple values"
            ],
            'date': [
                "Excel stores dates as numbers",
                "Use DATE function to create dates",
                "DATEDIF calculates date differences"
            ],
            'text': [
                "Use & operator for simple concatenation",
                "TRIM removes extra spaces",
                "SUBSTITUTE replaces text"
            ],
            'financial': [
                "Use absolute references ($) for constants",
                "PMT calculates loan payments",
                "FV calculates future value"
            ]
        }
        return tips.get(formula_type, ["Test your formula with sample data", "Use F9 to evaluate parts of formulas"])
    
    def clean_data(self, operations: List[str], sheet_name: str = None) -> Dict[str, Any]:
        """
        Clean and transform Excel data
        Operations can include:
        - 'remove_duplicates'
        - 'fill_missing_values'
        - 'convert_data_types'
        - 'normalize_text'
        - 'remove_outliers'
        - 'standardize_format'
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name and sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            if not sheet_name:
                sheet_name = list(self.sheets.keys())[0]
            
            df = self.sheets[sheet_name]['dataframe'].copy()
            original_shape = df.shape
            cleaning_report = {
                'operations_performed': [],
                'changes_made': {},
                'original_shape': original_shape,
                'final_shape': original_shape
            }
            
            for operation in operations:
                operation_lower = operation.lower()
                
                if 'remove_duplicates' in operation_lower:
                    before_count = len(df)
                    df = df.drop_duplicates()
                    after_count = len(df)
                    cleaning_report['operations_performed'].append('remove_duplicates')
                    cleaning_report['changes_made']['duplicates_removed'] = before_count - after_count
                
                elif 'fill_missing_values' in operation_lower:
                    missing_before = df.isnull().sum().sum()
                    # Fill numeric columns with mean
                    numeric_cols = df.select_dtypes(include=[np.number]).columns
                    for col in numeric_cols:
                        df[col] = df[col].fillna(df[col].mean())
                    
                    # Fill text columns with mode
                    text_cols = df.select_dtypes(include=['object']).columns
                    for col in text_cols:
                        mode_value = df[col].mode()[0] if len(df[col].mode()) > 0 else 'Unknown'
                        df[col] = df[col].fillna(mode_value)
                    
                    missing_after = df.isnull().sum().sum()
                    cleaning_report['operations_performed'].append('fill_missing_values')
                    cleaning_report['changes_made']['missing_values_filled'] = missing_before - missing_after
                
                elif 'convert_data_types' in operation_lower:
                    # Convert data types intelligently
                    for col in df.columns:
                        # Try to convert to numeric
                        try:
                            pd.to_numeric(df[col], errors='raise')
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                        except:
                            # Try to convert to datetime
                            try:
                                df[col] = pd.to_datetime(df[col], errors='coerce')
                            except:
                                # Keep as string
                                df[col] = df[col].astype(str)
                    
                    cleaning_report['operations_performed'].append('convert_data_types')
                    cleaning_report['changes_made']['data_types_converted'] = True
                
                elif 'normalize_text' in operation_lower:
                    text_cols = df.select_dtypes(include=['object']).columns
                    for col in text_cols:
                        # Remove extra spaces
                        df[col] = df[col].str.strip()
                        # Convert to title case
                        df[col] = df[col].str.title()
                    
                    cleaning_report['operations_performed'].append('normalize_text')
                    cleaning_report['changes_made']['text_normalized'] = len(text_cols)
                
                elif 'remove_outliers' in operation_lower:
                    numeric_cols = df.select_dtypes(include=[np.number]).columns
                    outliers_removed = 0
                    
                    for col in numeric_cols:
                        Q1 = df[col].quantile(0.25)
                        Q3 = df[col].quantile(0.75)
                        IQR = Q3 - Q1
                        lower_bound = Q1 - 1.5 * IQR
                        upper_bound = Q3 + 1.5 * IQR
                        
                        before_count = len(df)
                        df = df[(df[col] >= lower_bound) & (df[col] <= upper_bound)]
                        after_count = len(df)
                        outliers_removed += (before_count - after_count)
                    
                    cleaning_report['operations_performed'].append('remove_outliers')
                    cleaning_report['changes_made']['outliers_removed'] = outliers_removed
                
                elif 'standardize_format' in operation_lower:
                    # Standardize date formats
                    date_cols = df.select_dtypes(include=['datetime64']).columns
                    for col in date_cols:
                        df[col] = pd.to_datetime(df[col]).dt.strftime('%Y-%m-%d')
                    
                    # Standardize currency format
                    currency_pattern = r'\$[\d,]+\.?\d*'
                    for col in df.columns:
                        if df[col].dtype == 'object':
                            df[col] = df[col].astype(str).str.replace(currency_pattern, lambda x: f"${float(x.group().replace('$', '').replace(',', '')):,.2f}")
                    
                    cleaning_report['operations_performed'].append('standardize_format')
                    cleaning_report['changes_made']['format_standardized'] = True
            
            # Update the sheet with cleaned data
            self.sheets[sheet_name]['dataframe'] = df
            self.dataframes[sheet_name] = df
            cleaning_report['final_shape'] = df.shape
            
            return {
                'status': 'success',
                'cleaning_report': cleaning_report,
                'data_preview': df.head(10).to_dict('records'),
                'summary': f"Cleaned data: {cleaning_report['operations_performed']}"
            }
            
        except Exception as e:
            return {'error': f'Error cleaning data: {str(e)}'}
    
    def create_chart(self, chart_type: str, x_column: str, y_column: str, sheet_name: str = None, 
                    title: str = None, options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Create advanced charts from Excel data
        Chart types: bar, line, pie, scatter, area, histogram, boxplot
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name and sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            if not sheet_name:
                sheet_name = list(self.sheets.keys())[0]
            
            df = self.sheets[sheet_name]['dataframe']
            
            if x_column not in df.columns:
                return {'error': f'Column "{x_column}" not found'}
            if y_column not in df.columns:
                return {'error': f'Column "{y_column}" not found'}
            
            # Set up matplotlib
            plt.style.use('default')
            fig, ax = plt.subplots(figsize=(10, 6))
            
            chart_data = {
                'x': df[x_column].tolist(),
                'y': df[y_column].tolist(),
                'x_column': x_column,
                'y_column': y_column,
                'chart_type': chart_type
            }
            
            if chart_type.lower() == 'bar':
                ax.bar(df[x_column], df[y_column])
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Bar Chart: {y_column} by {x_column}')
            
            elif chart_type.lower() == 'line':
                ax.plot(df[x_column], df[y_column], marker='o')
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Line Chart: {y_column} over {x_column}')
            
            elif chart_type.lower() == 'pie':
                # For pie charts, we need to aggregate the data
                if pd.api.types.is_numeric_dtype(df[y_column]):
                    aggregated = df.groupby(x_column)[y_column].sum()
                    ax.pie(aggregated.values, labels=aggregated.index, autopct='%1.1f%%')
                else:
                    aggregated = df[x_column].value_counts()
                    ax.pie(aggregated.values, labels=aggregated.index, autopct='%1.1f%%')
                
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Pie Chart: {x_column} Distribution')
            
            elif chart_type.lower() == 'scatter':
                ax.scatter(df[x_column], df[y_column], alpha=0.6)
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Scatter Plot: {y_column} vs {x_column}')
            
            elif chart_type.lower() == 'area':
                ax.fill_between(df[x_column], df[y_column], alpha=0.6)
                ax.set_xlabel(x_column)
                ax.set_ylabel(y_column)
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Area Chart: {y_column} over {x_column}')
            
            elif chart_type.lower() == 'histogram':
                ax.hist(df[y_column], bins=20, alpha=0.7, edgecolor='black')
                ax.set_xlabel(y_column)
                ax.set_ylabel('Frequency')
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Histogram: {y_column} Distribution')
            
            elif chart_type.lower() == 'boxplot':
                ax.boxplot(df[y_column])
                ax.set_ylabel(y_column)
                if title:
                    ax.set_title(title)
                else:
                    ax.set_title(f'Box Plot: {y_column} Distribution')
            
            else:
                return {'error': f'Unsupported chart type: {chart_type}'}
            
            # Apply custom options
            if options:
                if 'color' in options:
                    ax.set_prop_cycle('color', [options['color']])
                if 'grid' in options and options['grid']:
                    ax.grid(True, alpha=0.3)
                if 'rotation' in options:
                    plt.xticks(rotation=options['rotation'])
            
            # Save chart to base64
            buffer = BytesIO()
            plt.tight_layout()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            chart_image = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
            
            # Store chart data
            chart_id = str(uuid.uuid4())
            self.charts[chart_id] = {
                'data': chart_data,
                'image': chart_image,
                'created_at': datetime.now().isoformat()
            }
            
            return {
                'status': 'success',
                'chart_id': chart_id,
                'chart_type': chart_type,
                'image': chart_image,
                'data_summary': {
                    'x_column': x_column,
                    'y_column': y_column,
                    'data_points': len(df),
                    'x_range': [df[x_column].min(), df[x_column].max()],
                    'y_range': [df[y_column].min(), df[y_column].max()]
                }
            }
            
        except Exception as e:
            return {'error': f'Error creating chart: {str(e)}'}
    
    def create_pivot_table(self, sheet_name: str, index_cols: List[str], 
                          value_cols: List[str], agg_func: str = 'sum',
                          filters: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Create pivot tables from Excel data
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            df = self.sheets[sheet_name]['dataframe']
            
            # Validate columns
            for col in index_cols + value_cols:
                if col not in df.columns:
                    return {'error': f'Column "{col}" not found'}
            
            # Apply filters if provided
            if filters:
                for col, condition in filters.items():
                    if col in df.columns:
                        if isinstance(condition, dict):
                            if 'min' in condition:
                                df = df[df[col] >= condition['min']]
                            if 'max' in condition:
                                df = df[df[col] <= condition['max']]
                            if 'values' in condition:
                                df = df[df[col].isin(condition['values'])]
                        else:
                            df = df[df[col] == condition]
            
            # Create pivot table
            pivot_table = pd.pivot_table(
                df,
                index=index_cols,
                values=value_cols,
                aggfunc=agg_func,
                fill_value=0
            )
            
            # Generate pivot table ID
            pivot_id = str(uuid.uuid4())
            self.pivot_tables[pivot_id] = {
                'data': pivot_table.to_dict(),
                'index_columns': index_cols,
                'value_columns': value_cols,
                'aggregation': agg_func,
                'filters': filters,
                'created_at': datetime.now().isoformat()
            }
            
            return {
                'status': 'success',
                'pivot_id': pivot_id,
                'pivot_table': pivot_table.to_dict(),
                'summary': {
                    'index_columns': index_cols,
                    'value_columns': value_cols,
                    'aggregation': agg_func,
                    'total_rows': len(pivot_table),
                    'total_columns': len(pivot_table.columns)
                },
                'data_preview': pivot_table.head(10).to_dict()
            }
            
        except Exception as e:
            return {'error': f'Error creating pivot table: {str(e)}'}
    
    def validate_data(self, sheet_name: str, validation_rules: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate data against custom rules
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            df = self.sheets[sheet_name]['dataframe']
            validation_results = {
                'passed': True,
                'errors': [],
                'warnings': [],
                'summary': {}
            }
            
            for column, rules in validation_rules.items():
                if column not in df.columns:
                    validation_results['errors'].append(f'Column "{column}" not found')
                    validation_results['passed'] = False
                    continue
                
                col_validation = {
                    'column': column,
                    'total_rows': len(df),
                    'valid_rows': 0,
                    'invalid_rows': 0,
                    'errors': []
                }
                
                # Apply validation rules
                for rule_type, rule_value in rules.items():
                    if rule_type == 'required':
                        if rule_value:
                            null_count = df[column].isnull().sum()
                            if null_count > 0:
                                col_validation['errors'].append(f'{null_count} null values found')
                                col_validation['invalid_rows'] += null_count
                    
                    elif rule_type == 'min':
                        if pd.api.types.is_numeric_dtype(df[column]):
                            below_min = (df[column] < rule_value).sum()
                            if below_min > 0:
                                col_validation['errors'].append(f'{below_min} values below minimum {rule_value}')
                                col_validation['invalid_rows'] += below_min
                    
                    elif rule_type == 'max':
                        if pd.api.types.is_numeric_dtype(df[column]):
                            above_max = (df[column] > rule_value).sum()
                            if above_max > 0:
                                col_validation['errors'].append(f'{above_max} values above maximum {rule_value}')
                                col_validation['invalid_rows'] += above_max
                    
                    elif rule_type == 'pattern':
                        if df[column].dtype == 'object':
                            pattern = re.compile(rule_value)
                            invalid_pattern = df[column].astype(str).apply(lambda x: not pattern.match(x)).sum()
                            if invalid_pattern > 0:
                                col_validation['errors'].append(f'{invalid_pattern} values don\'t match pattern {rule_value}')
                                col_validation['invalid_rows'] += invalid_pattern
                    
                    elif rule_type == 'unique':
                        if rule_value:
                            duplicates = df[column].duplicated().sum()
                            if duplicates > 0:
                                col_validation['errors'].append(f'{duplicates} duplicate values found')
                                col_validation['invalid_rows'] += duplicates
                    
                    elif rule_type == 'in_range':
                        if pd.api.types.is_numeric_dtype(df[column]):
                            out_of_range = ((df[column] < rule_value[0]) | (df[column] > rule_value[1])).sum()
                            if out_of_range > 0:
                                col_validation['errors'].append(f'{out_of_range} values out of range {rule_value}')
                                col_validation['invalid_rows'] += out_of_range
                
                col_validation['valid_rows'] = col_validation['total_rows'] - col_validation['invalid_rows']
                
                if col_validation['errors']:
                    validation_results['errors'].extend(col_validation['errors'])
                    validation_results['passed'] = False
                
                validation_results['summary'][column] = col_validation
            
            return validation_results
            
        except Exception as e:
            return {'error': f'Error validating data: {str(e)}'}
    
    def generate_automation_script(self, task_description: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        Generate Python/Pandas automation scripts
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name and sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            if not sheet_name:
                sheet_name = list(self.sheets.keys())[0]
            
            df = self.sheets[sheet_name]['dataframe']
            
            # Generate script based on task description
            script_template = f"""
import pandas as pd
import numpy as np
from datetime import datetime

# Load the Excel file
df = pd.read_excel('your_file.xlsx', sheet_name='{sheet_name}')

# Task: {task_description}

# Your automation code here:
# Example operations:
# 1. Data cleaning
# df = df.dropna()  # Remove rows with missing values
# df = df.drop_duplicates()  # Remove duplicate rows

# 2. Data transformation
# df['new_column'] = df['existing_column'] * 2

# 3. Data analysis
# summary = df.groupby('category')['value'].sum()

# 4. Export results
# df.to_excel('processed_data.xlsx', index=False)
# summary.to_excel('summary.xlsx')

print("Automation completed successfully!")
"""
            
            script_id = str(uuid.uuid4())
            self.automation_scripts[script_id] = {
                'task': task_description,
                'script': script_template,
                'sheet_name': sheet_name,
                'created_at': datetime.now().isoformat()
            }
            
            return {
                'status': 'success',
                'script_id': script_id,
                'script': script_template,
                'task': task_description,
                'recommendations': [
                    'Customize the script based on your specific needs',
                    'Add error handling for robust automation',
                    'Test the script on a copy of your data first',
                    'Consider adding logging for tracking progress'
                ]
            }
            
        except Exception as e:
            return {'error': f'Error generating automation script: {str(e)}'}
    
    def export_data(self, format_type: str, sheet_name: str = None, 
                   filters: Dict[str, Any] = None, filename: str = None) -> Dict[str, Any]:
        """
        Export data in various formats
        """
        try:
            if not self.sheets:
                return {'error': 'No Excel file loaded'}
            
            if sheet_name and sheet_name not in self.sheets:
                return {'error': f'Sheet "{sheet_name}" not found'}
            
            if not sheet_name:
                sheet_name = list(self.sheets.keys())[0]
            
            df = self.sheets[sheet_name]['dataframe'].copy()
            
            # Apply filters if provided
            if filters:
                for col, condition in filters.items():
                    if col in df.columns:
                        if isinstance(condition, dict):
                            if 'min' in condition:
                                df = df[df[col] >= condition['min']]
                            if 'max' in condition:
                                df = df[df[col] <= condition['max']]
                            if 'values' in condition:
                                df = df[df[col].isin(condition['values'])]
                        else:
                            df = df[df[col] == condition]
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"export_{sheet_name}_{timestamp}"
            
            export_data = {}
            
            if format_type.lower() == 'csv':
                csv_data = df.to_csv(index=False)
                export_data['csv'] = csv_data
                export_data['filename'] = f"{filename}.csv"
            
            elif format_type.lower() == 'json':
                json_data = df.to_json(orient='records', indent=2)
                export_data['json'] = json_data
                export_data['filename'] = f"{filename}.json"
            
            elif format_type.lower() == 'excel':
                # Create Excel file with formatting
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                
                # Write data
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # Add formatting
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to buffer
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                excel_data = base64.b64encode(buffer.getvalue()).decode()
                
                export_data['excel'] = excel_data
                export_data['filename'] = f"{filename}.xlsx"
            
            elif format_type.lower() == 'html':
                html_data = df.to_html(index=False, classes='table table-striped')
                export_data['html'] = html_data
                export_data['filename'] = f"{filename}.html"
            
            else:
                return {'error': f'Unsupported export format: {format_type}'}
            
            return {
                'status': 'success',
                'format': format_type,
                'data': export_data,
                'summary': {
                    'rows_exported': len(df),
                    'columns_exported': len(df.columns),
                    'filters_applied': bool(filters)
                }
            }
            
        except Exception as e:
            return {'error': f'Error exporting data: {str(e)}'}

# Create a global instance
excel_bot = AdvancedExcelChatbot()

def excel_answer(file_path: str, question: str) -> str:
    """Legacy function for backward compatibility"""
    try:
        # Load file if not already loaded
        if not excel_bot.sheets:
            result = excel_bot.load_excel_file(file_path)
            if result['status'] == 'error':
                return f"Error loading file: {result['message']}"
        
        # Process the question
        result = excel_bot.natural_language_query(question)
        
        if 'error' in result:
            return f"Error: {result['error']}"
        
        # Format the response
        if result['type'] == 'aggregation':
            return result['formatted_results']
        elif result['type'] == 'filter':
            return f"Found {result['filtered_rows']} rows matching your criteria. Showing first 20 results."
        elif result['type'] == 'groupby':
            return result['summary']
        elif result['type'] == 'top_n':
            return result['summary']
        else:
            return result.get('message', 'Query processed successfully')
            
    except Exception as e:
        return f"Error processing question: {str(e)}"
